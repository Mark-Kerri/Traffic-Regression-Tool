import os
from copy import copy
from calendar import month_abbr
import openpyxl


# Constants
TEMPLATE_PATH = "data/utils/excel_template_v0.01.xlsx"
OUTPUT_DIR = "data/reg_input"
HEADER_STYLE_CELL = "D13"
VALUE_STYLE_CELL = "D14"


def copy_cell_style(source_cell, target_cell):
    # Use a try-except block to handle potential attribute errors
    try:
        # Copy font
        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        # Copy border
        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        # Copy fill
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)

        # Copy number format
        target_cell.number_format = source_cell.number_format

        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        # Copy protection
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    except AttributeError as e:
        print(f"Warning: Unable to copy some cell properties. Error: {e}")


def generate_timeline(timeline_inputs):
    timestep = timeline_inputs["Timestep"]
    start_year = timeline_inputs["Start Year"]
    start_step = timeline_inputs["Start Timestep"]
    end_year = timeline_inputs["End Year"]
    end_step = timeline_inputs["End Timestep"]

    # Validation
    if start_year > end_year:
        raise ValueError("Start Year must be less than or equal to End Year")

    if timestep not in ["Monthly", "Quarterly", "Yearly"]:
        raise ValueError("Timestep must be Monthly, Quarterly, or Yearly")

    if timestep == "Monthly" and (
        start_step < 1 or start_step > 12 or end_step < 1 or end_step > 12
    ):
        raise ValueError(
            "For Monthly timestep, Start and End Timestep must be between 1 and 12"
        )
    elif timestep == "Quarterly" and (
        start_step < 1 or start_step > 4 or end_step < 1 or end_step > 4
    ):
        raise ValueError(
            "For Quarterly timestep, Start and End Timestep must be between 1 and 4"
        )
    elif timestep == "Yearly" and (start_step != 1 or end_step != 1):
        raise ValueError("For Yearly timestep, Start and End Timestep must be 1")

    list_1, list_2, list_3 = [], [], []

    steps_per_year = (
        12 if timestep == "Monthly" else 4 if timestep == "Quarterly" else 1
    )

    for year in range(start_year, end_year + 1):
        start = start_step if year == start_year else 1
        end = end_step if year == end_year else steps_per_year

        for step in range(start, end + 1):
            list_1.append(year)

            if timestep == "Monthly":
                step_str = month_abbr[step]
            elif timestep == "Quarterly":
                step_str = f"Q{step}"
            else:  # Yearly
                step_str = ""

            list_2.append(step_str)
            list_3.append(f"{year} {step_str}".strip())

    # Create a dictionary to hold the timeline data:
    # 'years': List of years for each timestep
    # 'steps': List of timestep labels (e.g., 'Q1', 'Jan', etc.)
    # 'combined': List of combined year and step (e.g., '2022 Q1')
    timelines = {"years": list_1, "steps": list_2, "combined": list_3}

    return timelines


def create_input_template(name_variables, y_variables, x_variables, timeline_inputs,file_name,output_folder_path):
    try:
        # Load the template workbook
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active

        # Update basic information
        client_name = name_variables["Client"]
        project_name = name_variables["Project"]
        # file_name = f"{project_name} Regression Inputs"

        update_basic_info(ws, client_name, project_name, file_name)

        # Generate timeline
        timelines = generate_timeline(timeline_inputs)

        # Insert columns before adding any timelines, maintaining styles from column G
        num_columns = len(timelines["combined"])
        insert_columns_with_style(ws, 7, num_columns)

        # Add variables
        last_row = add_variables_with_timeline(
            ws, y_variables, "Dependent Variables", 13, timelines
        )
        add_variables_with_timeline(
            ws, x_variables, "Independent Variables", last_row + 5, timelines
        )

        # Save the modified workbook

        output_path = os.path.join(output_folder_path, f"{file_name}.xlsx")
        wb.save(output_path)
        wb.close()

        print(f"Template created successfully: {output_path}")

    except Exception as e:
        print(f"Error creating template: {e}")


def insert_columns_with_style(ws, start_column, num_columns):
    # Get the maximum row in the worksheet
    max_row = ws.max_row

    # Insert new columns
    ws.insert_cols(start_column, num_columns)

    # Copy styles from the original column G to the new columns
    for row in range(1, max_row + 1):
        source_cell = ws.cell(row=row, column=start_column + num_columns)
        for col in range(start_column, start_column + num_columns):
            target_cell = ws.cell(row=row, column=col)
            copy_cell_style(source_cell, target_cell)


def update_basic_info(ws, client_name, project_name, file_name):
    ws["B2"] = client_name
    ws["B3"] = project_name
    ws["B4"] = file_name
    ws["B6"] = "Regression Inputs"


def add_variables_with_timeline(ws, variables, header_text, start_row, timelines):
    ws[f"D{start_row}"] = header_text
    ws[f"E{start_row}"] = "abs/pct"

    header_style = ws[HEADER_STYLE_CELL]
    value_style = ws[VALUE_STYLE_CELL]
    year_step_style = ws["D12"]

    copy_cell_style(header_style, ws[f"D{start_row}"])
    copy_cell_style(header_style, ws[f"E{start_row}"])

    # Add year list
    for col, value in enumerate(timelines["years"], start=7):
        cell = ws.cell(row=start_row - 2, column=col)
        cell.value = value
        copy_cell_style(year_step_style, cell)

    # Add step list
    for col, value in enumerate(timelines["steps"], start=7):
        cell = ws.cell(row=start_row - 1, column=col)
        cell.value = value
        copy_cell_style(year_step_style, cell)

    # Add combined timeline headers
    for col, value in enumerate(timelines["combined"], start=7):
        cell = ws.cell(row=start_row, column=col)
        cell.value = value
        copy_cell_style(header_style, cell)

    row = start_row
    for key, item in variables.items():
        row += 1
        ws[f"D{row}"] = key
        ws[f"E{row}"] = item

        copy_cell_style(value_style, ws[f"D{row}"])
        copy_cell_style(value_style, ws[f"E{row}"])

        # Add empty cells for timeline values
        for col in range(7, 7 + len(timelines["combined"])):
            cell = ws.cell(row=row, column=col)
            copy_cell_style(value_style, cell)

    return row


# if __name__ == "__main__":
#     name_variables = {"Client": "AEDL", "Project": "A32 LV Traffic Forecast"}
#     y_variables = {"Traffic (AADT)": "abs"}
#     x_variables = {
#         "GDP": "abs",
#         "Unemployment": "pct",
#         "Ramp Up": "abs",
#         "Signage": "abs",
#         "Cong_IC2": "abs",
#         "Cong_A20": "abs",
#     }
#     timeline_inputs = {
#         "Timestep": "Quarterly",
#         "Start Year": 2012,
#         "Start Timestep": 1,
#         "End Year": 2023,
#         "End Timestep": 4,
#     }
#     create_input_template(name_variables, y_variables, x_variables, timeline_inputs)
