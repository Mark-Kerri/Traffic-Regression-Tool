import os
import openpyxl
import pandas as pd

# Assumptions for the following script to work:
# Column D only has values for dependent and independent variables
# and no other content (otherwise any column D text will be assumed a new variable)
# Data for each variable found in Column D is being read from column G onwards
# until an empty cell is found (None)
# Output dataframe will name each variable before "Independent Variables" cell
# as X:{variable name} and every variable after "Independent Variables" cell as Y:{variable name}
# Building the index of the df assuming column G is empty (on all rows above data)
# and column G is the first time series column


def spreadsheet_to_df(file_name):

    # delete below line to test pathing
    file_path = os.path.join("data", "reg_input")
    file_name = "Development Test Data Regression Inputs.xlsx"
    full_path = os.path.join(file_path, file_name)
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    sheet = workbook.active

    # Read column names and row names so that the dataframe is filled
    df_cols = []
    dependent_flag = 1
    col = 4  # start from Column D for variable names
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is not None:
            # if sheet.cell(row=row, column=col).value != 'Dependent Variables'
            # and  sheet.cell(row=row, column=col).value != 'Independent Variables' :
            if sheet.cell(row=row, column=col).value == "Dependent Variables":
                dependent_flag = 1
                continue
            if sheet.cell(row=row, column=col).value == "Independent Variables":
                dependent_flag = 0
                continue
            if dependent_flag == 1:
                df_cols.append("y:" + sheet.cell(row=row, column=col).value)
            elif dependent_flag == 0:
                df_cols.append("x:" + sheet.cell(row=row, column=col).value)

    # Build dataframe by iterating over the columns for each variable
    df = pd.DataFrame(columns=df_cols)
    for c in df_cols:
        temp_list = []
        # print(c)
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=col).value == c[2:]:
                start_row = row
                break
        cell_names = sheet[start_row]
        for cell_obj in cell_names[6:]:
            try:
                val = float(str(cell_obj.value))
                temp_list.append(val)
            except ValueError:
                break
        df[c] = temp_list

    # Build the index of the df assuming column G is empty
    # and column G is the first time series column
    col = 7  # Start at Col G
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is not None:
            start_row = row
            print(start_row)
            break
    idx_row = start_row + 2  # 2 rows below first non-empty cell

    temp_list = []
    cell_names = sheet[idx_row]
    for cell_obj in cell_names[(col - 1) :]:
        try:
            val = str(cell_obj.value)
            if val != "None":
                temp_list.append(val)
        except ValueError:
            break
    df_idx = temp_list
    df.index = df_idx

    return df
