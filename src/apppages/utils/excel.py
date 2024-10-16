"""
Excel Input Template Generator Module.

This module provides functionality for generating and managing Excel input
templates for project data analysis. It uses the `openpyxl` library to create,
manipulate, and style Excel spreadsheets based on various project inputs and
timeline details.

Constants:
- TEMPLATE_PATH (str): Path to the Excel template file used as the basis for
  the generated input files.
- OUTPUT_DIR (str): Directory path where the output files will be saved.
- HEADER_STYLE_CELL (str): Cell reference for the header style used in the
  generated template.
- VALUE_STYLE_CELL (str): Cell reference for the value style used in the
  generated template.

Functions:
- copy_cell_style(source_cell, target_cell):
    Copies the style attributes from a source cell to a target cell, including
    font, border, fill, number format, alignment, and protection.

- generate_timeline(timeline_inputs):
    Generates a timeline based on input parameters such as timestep, start year,
    and end year. Returns a dictionary with years, steps, and combined timeline
    strings.

- create_input_template(name_variables, y_variables, x_variables,
    timeline_inputs, file_name, output_folder_path):
    Creates an Excel input template based on provided project details, variables,
    and timeline inputs. Handles exceptions related to file operations and
    input validation, saving the final template to the specified directory.

- insert_columns_with_style(ws, start_column, num_columns):
    Inserts columns into the specified worksheet and applies the style from an
    existing column to the new columns.

- update_basic_info(ws, client_name, project_name, file_name):
    Updates the basic project information (client name, project name, and file
    name) in the worksheet.

- add_variables_with_timeline(ws, variables, header_text, start_row, timelines):
    Adds variable names, types, and corresponding timeline data to the worksheet,
    including the header and value styling.

This module is designed to facilitate the creation of structured Excel templates
for the tool, ensuring consistency and accuracy in the data analysis.
"""

import os
import io
from copy import copy
from calendar import month_abbr
import xlsxwriter
import openpyxl
import pandas as pd
from datetime import datetime
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

# Constants
TEMPLATE_PATH = "data/utils/excel_template_v0.01.xlsx"
OUTPUT_DIR = "data/reg_input"
HEADER_STYLE_CELL = "D13"
VALUE_STYLE_CELL = "D14"


def copy_cell_style(source_cell, target_cell):
    """
    Copy the style of a source cell to a target cell.

    Parameters:
        source_cell (openpyxl.cell.Cell): The cell from which to copy the style.
        target_cell (openpyxl.cell.Cell): The cell to which the style is copied.
    """
    try:
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
    except AttributeError as e:
        print(f"Warning: Unable to copy some cell properties. Error: {e}")


def generate_timeline(timeline_inputs):
    """
    Generate a timeline based on the provided inputs.

    Parameters:
        timeline_inputs (dict): A dictionary containing timeline parameters
                                such as 'Timestep', 'Start Year', 'End Year', etc.

    Returns:
        dict: A dictionary with keys 'years', 'steps', and 'combined',
              representing different components of the timeline.
    """
    timestep = timeline_inputs["Timestep"]
    start_year = timeline_inputs["Start Year"]
    start_step = timeline_inputs["Start Timestep"]
    end_year = timeline_inputs["End Year"]
    end_step = timeline_inputs["End Timestep"]

    # Validation
    if start_year > end_year:
        raise ValueError("Start Year must be less than or equal to End Year")

    if timestep not in ["Monthly", "Quarterly", "Yearly"]:
        raise ValueError("Timestep must be Monthly, Quarterly, or Yearly")

    if timestep == "Monthly" and (
        start_step < 1 or start_step > 12 or end_step < 1 or end_step > 12
    ):
        raise ValueError(
            "For Monthly timestep, Start and End Timestep must be between 1 and 12"
        )
    elif timestep == "Quarterly" and (
        start_step < 1 or start_step > 4 or end_step < 1 or end_step > 4
    ):
        raise ValueError(
            "For Quarterly timestep, Start and End Timestep must be between 1 and 4"
        )
    elif timestep == "Yearly" and (start_step != 1 or end_step != 1):
        raise ValueError("For Yearly timestep, Start and End Timestep must be 1")

    list_1, list_2, list_3 = [], [], []
    steps_per_year = (
        12 if timestep == "Monthly" else 4 if timestep == "Quarterly" else 1
    )

    for year in range(start_year, end_year + 1):
        start = start_step if year == start_year else 1
        end = end_step if year == end_year else steps_per_year

        for step in range(start, end + 1):
            list_1.append(year)

            if timestep == "Monthly":
                step_str = month_abbr[step]
            elif timestep == "Quarterly":
                step_str = f"Q{step}"

            else:  # Yearly
                step_str = ""

            list_2.append(step_str)
            list_3.append(f"{year} {step_str}".strip())

    timelines = {"years": list_1, "steps": list_2, "combined": list_3}#,"seasonality": list_4}
    return timelines


def create_input_template(
    name_variables,
    y_variables,
    x_variables,
    timeline_inputs,
    file_name,
    output_folder_path,
):
    """
    Create an Excel input template based on project details, variables, and timeline inputs.

    Parameters:
        name_variables (dict): Project and client names.
        y_variables (dict): Dependent variables and their types.
        x_variables (dict): Independent variables and their types.
        timeline_inputs (dict): Timeline details. Saves the timestep on cell H50 (see timeline_inputs["Timestep"] )
        file_name (str): Desired name for the output file.
        output_folder_path (str): Directory path where the file will be saved.

    Raises:
        FileNotFoundError: If the template file is not found.
        KeyError: If a required key is missing in input dictionaries.
        TypeError: If there is a type mismatch in the inputs.
        openpyxl.utils.exceptions.InvalidFileException: If the Excel template is invalid.
        OSError: If there is a problem with file I/O operations.
    """
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active

        client_name = name_variables["Client"]
        project_name = name_variables["Project"]
        update_basic_info(ws, client_name, project_name, file_name)

        timelines = generate_timeline(timeline_inputs)
        num_columns = len(timelines["combined"])
        insert_columns_with_style(ws, 7, num_columns)

        last_row = add_variables_with_timeline(
            ws, y_variables, "Dependent Variables", 13, timelines
        )
        add_variables_with_timeline(
            ws, x_variables, "Independent Variables", last_row + 5, timelines
        )
        ws["H50"] = timeline_inputs["Timestep"]
        output_path = os.path.join(output_folder_path, f"{file_name}.xlsx")
        # wb.save(output_path)
        # wb.close()

        # fill in seasonality timelines here:
        # find row number where "Seasonality" is mentioned on col D
        # then iterate across columns and fill with 1s the rows where "Q1 " or "Jan" (First three characters?)
        # match with the variable name
        # workbook = openpyxl.load_workbook(input_file_path, data_only=True)
        # sheet = workbook.active
        var_col = 4  # start from Column D for variable names
        seas_var_dict = {}
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=var_col).value
            if cell_value and 'Dependent Variable' in cell_value:
                header_row = row
                # print(header_row)
            if cell_value and 'Seasonality' in cell_value:
                seas_row_num = row
                seas_var_ref = cell_value.split(' ')[0]
                seas_var_dict[seas_row_num] = seas_var_ref
                # print(seas_var_dict)


        for seas_row,seas_ref in seas_var_dict.items():
            for col in range(var_col+3,var_col+3+len(timelines["combined"])):
                header_cell_value = ws.cell(row=header_row, column=col).value
                if header_cell_value and seas_ref in header_cell_value:
                    ws.cell(row=seas_row, column=col).value = 1
                elif header_cell_value and seas_ref not in header_cell_value:
                    ws.cell(row=seas_row, column=col).value = 0
            # print(ws.cell(row=seas_row_num, column=col).value)
        # Save the workbook to an in-memory buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)  # Reset buffer pointer to the beginning
        wb.close()

        print(f"Template created successfully: {output_path}")
        return buffer

    except FileNotFoundError:
        print(f"Template file not found at path: {TEMPLATE_PATH}")
    except KeyError as ke:
        print(f"Missing key in input data: {ke}")
    except TypeError as te:
        print(f"Type error encountered: {te}")
    except openpyxl.utils.exceptions.InvalidFileException as ife:
        print(f"Invalid file format encountered: {ife}")
    except OSError as oe:
        print(f"File system error when saving the file: {oe}")


def insert_columns_with_style(ws, start_column, num_columns):
    """
    Insert columns into the worksheet and applies the style from an existing column.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to modify.
        start_column (int): The column index where insertion starts.
        num_columns (int): Number of columns to insert.
    """
    max_row = ws.max_row
    ws.insert_cols(start_column, num_columns)

    for row in range(1, max_row + 1):
        source_cell = ws.cell(row=row, column=start_column + num_columns)
        for col in range(start_column, start_column + num_columns):
            target_cell = ws.cell(row=row, column=col)
            copy_cell_style(source_cell, target_cell)


def update_basic_info(ws, client_name, project_name, file_name):
    """
    Update basic project information in the worksheet.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to update.
        client_name (str): Name of the client.
        project_name (str): Name of the project.
        file_name (str): The name of the output file.
    """
    ws["B2"] = client_name
    ws["B3"] = project_name
    ws["B4"] = file_name
    ws["B6"] = "Regression Inputs"


def add_variables_with_timeline(ws, variables, header_text, start_row, timelines):
    """
    Add variables and their corresponding timeline to the worksheet.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to update.
        variables (dict): A dictionary of variables and their types.
        header_text (str): The header text to display.
        start_row (int): The starting row for adding the variables.
        timelines (dict): The generated timeline data.

    Returns:
        int: The last row index after adding the variables.
    """
    ws[f"D{start_row}"] = header_text
    ws[f"E{start_row}"] = "abs/pct"

    header_style = ws[HEADER_STYLE_CELL]
    value_style = ws[VALUE_STYLE_CELL]
    year_step_style = ws["D12"]

    copy_cell_style(header_style, ws[f"D{start_row}"])
    copy_cell_style(header_style, ws[f"E{start_row}"])

    for col, value in enumerate(timelines["years"], start=7):
        cell = ws.cell(row=start_row - 2, column=col)
        cell.value = value
        copy_cell_style(year_step_style, cell)

    for col, value in enumerate(timelines["steps"], start=7):
        cell = ws.cell(row=start_row - 1, column=col)
        cell.value = value
        copy_cell_style(year_step_style, cell)

    for col, value in enumerate(timelines["combined"], start=7):
        cell = ws.cell(row=start_row, column=col)
        cell.value = value
        copy_cell_style(header_style, cell)

    row = start_row
    for key, item in variables.items():
        row += 1
        ws[f"D{row}"] = key
        ws[f"E{row}"] = item

        copy_cell_style(value_style, ws[f"D{row}"])
        copy_cell_style(value_style, ws[f"E{row}"])

        for col in range(7, 7 + len(timelines["combined"])):
            cell = ws.cell(row=row, column=col)
            copy_cell_style(value_style, cell)

    return row


def spreadsheet_to_df(input_file_path):
    """
    Convert data in the Excel spreadsheet temlate into a pandas DataFrame.

    This function reads an Excel file, processes specific columns and rows to extract dependent
    and independent variables data, and then constructs a DataFrame. It also identifies the
    time series index and gathers variable type/unit information into a dictionary.

    Assumptions for the following script to work:
    Column D only has values for dependent and independent variables
    and no other content (otherwise any column D text will be assumed a new variable)
    Data for each variable found in Column D is being read from column G onwards
    until an empty cell is found (None)
    Output dataframe will name each variable before "Independent Variables" cell
    as X:{variable name} and every variable after "Independent Variables" cell as Y:{variable name}
    Building the index of the df assuming column G is empty (on all rows above data)
    and column G is the first time series column

    Parameters:
        input_file_path (str): The file path to the input Excel file.

    Returns:
        tuple: A tuple containing:
            - df (pd.DataFrame): The resulting DataFrame with dependent and independent variables.
            - df_index (list): A list representing the DataFrame's time series index.
            - var_dict (dict): A dictionary mapping variables to their unit type (e.g. "abs/pct").
    """
    workbook = openpyxl.load_workbook(input_file_path, data_only=True)
    sheet = workbook.active

    # Read column names and row names so that the dataframe is filled
    df_cols = []
    dependent_flag = 1
    col = 4  # start from Column D for variable names
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is not None:
            if sheet.cell(row=row, column=col).value == "Dependent Variables":
                dependent_flag = 1
                continue
            if sheet.cell(row=row, column=col).value == "Independent Variables":
                dependent_flag = 0
                continue
            if dependent_flag == 1:
                df_cols.append("y:" + sheet.cell(row=row, column=col).value)
            elif dependent_flag == 0:
                df_cols.append("x:" + sheet.cell(row=row, column=col).value)

    # Build dataframe by iterating over the columns for each variable
    df = pd.DataFrame(columns=df_cols)
    for c in df_cols:
        temp_list = []
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=col).value == c[2:]:
                start_row = row
                break
        cell_names = sheet[start_row]
        for cell_obj in cell_names[6:]:
            try:
                val = float(str(cell_obj.value))
                temp_list.append(val)
            except ValueError:
                break
        df[c] = temp_list

    # Build the index of the df assuming column G is empty
    # and column G is the first time series column
    col = 7  # Start at Col G
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is not None:
            start_row = row
            # print(start_row)
            break
    idx_row = start_row + 2  # 2 rows below first non-empty cell

    temp_list = []
    cell_names = sheet[idx_row]
    for cell_obj in cell_names[(col - 1) :]:
        try:
            val = str(cell_obj.value)
            if val != "None":
                temp_list.append(val)
        except ValueError:
            break
    df_idx = temp_list
    df.index = df_idx

    # Read variable type/unit and build a dictionary
    var_dict = {}
    col = 5  # Start at Column E
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value == "abs/pct" or (
            sheet.cell(row=row, column=col).value is None
        ):
            continue
        var_dict[sheet.cell(row=row, column=col - 1).value] = sheet.cell(
            row=row, column=col
        ).value
    df_index = df.index

    # Read timestep from cell H 50
    # timestep = sheet.cell(row=50, column=8).value
    row_11_numbers = [cell.value for cell in sheet[11]]  # Row indexing in openpyxl starts from 1
    my_series = pd.Series(row_11_numbers)
    counts = my_series.value_counts(dropna=True)
    prd = counts.max() ### check this
    if prd == 4:
        timestep = "Quarterly"
    elif prd == 12:
        timestep = "Monthly"
    elif prd == 1:
        timestep = "Yearly"
    return df, df_index, var_dict, timestep


# def export_to_excel(coeff_df,df,summary_df,residuals_df,path):


def export_to_excel(
    regressions_df,
    g_df,
    test_name,
    coeff_df,
    summary_df,
    residuals_df,
    path,
    forecast_df,
):

    # # avoid having too long worksheet names, which causes errors when saving workbooks (max chars 31)
    # workbook_test_name = test_name
    # if len(test_name) > 20:
    #     shortened_test_name_elements = [x[:3] for x in test_name.split("-")]
    #     workbook_test_name = "".join(shortened_test_name_elements)
    #     if len(workbook_test_name) > 20:
    #         workbook_test_name = workbook_test_name[:15] + "+"
    workbook_test_name = shorten_test_name(test_name)
    current_timestamp = datetime.now().strftime("%Y-%m-%d-%H%M")
    cols = [x for x in forecast_df.columns if x.startswith("y:") or x == "Forecast y"]
    forecast_df = forecast_df[cols]
    forecast_df_cols = len(forecast_df.columns)
    full_output_path = os.path.join(
        path, f"{current_timestamp}_{workbook_test_name}_output.xlsx"
    )

    # Create an in-memory buffer to hold the Excel file
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:

        # all regressions
        regressions_df.to_excel(writer, sheet_name=f"All regressions", index=True)
        ws_regressions = writer.sheets[f"All regressions"]
        ws_regressions = ws_regressions.set_column(0, 0, 85)
        # Define styles for bold text and grey background
        bold_font = Font(bold=True)
        grey_fill = PatternFill(
            start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
        )

        # add metadata
        # metadata = pd.DataFrame(data=base_year_datapoints)
        # metadata.to_excel(writer, sheet_name=f'Base year', index=False)

        # Write each dataframe to a different worksheet.
        coeff_df[~coeff_df.index.str.contains("t-val")].to_excel(writer, sheet_name=f"{workbook_test_name} Coeffs", index=True)
        summary_tables = pd.DataFrame()
        for table in summary_df.tables:
            summary_tables = pd.concat([summary_tables, pd.DataFrame(table)])
            summary_tables.to_excel(
                writer, sheet_name=f"{workbook_test_name} Summ", index=False
            )
        residuals_df.to_excel(
            writer, sheet_name=f"{workbook_test_name} Rsdl", index=True
        )

        ws_coeff = writer.sheets[f"{workbook_test_name} Coeffs"]
        ws_coeff = ws_coeff.set_column(0, 0, 25)

        workbook = writer.book
        worksheet = writer.sheets[f"{workbook_test_name} Rsdl"]

        # Define the chart object (scatter plot)
        chart = workbook.add_chart({"type": "scatter"})

        # Add multiple series to the scatter plot
        for i in range(1, 4):  # Assuming 3 columns of Y-values
            chart.add_series(
                {
                    "name": [f"{workbook_test_name} Rsdl", 0, i],  # Column header as series name
                    "categories": [f"{workbook_test_name} Rsdl", 1, 0, len(residuals_df), 0],  # X values (Column A)
                    "values": [f"{workbook_test_name} Rsdl", 1, i, len(residuals_df), i],  # Y values (Columns B, C, D)
                    "marker": {"type": "circle", "size": 5},
                }
            )

        # Set chart title and labels
        chart.set_title({"name": "Residuals Scatter Plot"})
        chart.set_x_axis({"name": "Time",
                          # "label_position": "low",  # Set labels to appear low on the axis
                          # "visible": True,  # Ensure the axis is visible
                          })  # X-axis label
        chart.set_y_axis({"name": "Residuals"})  # Y-axis label

        # Insert the chart into the worksheet
        worksheet.insert_chart("F2", chart)

        forecast_df.to_excel(
            writer, sheet_name=f"{workbook_test_name} Predicted", index=True
        )

        worksheet = writer.sheets[f"{workbook_test_name} Predicted"]

        # Define the chart object
        chart = workbook.add_chart({"type": "line"})

        # Add the first series (Column B as Y values)
        # chart.add_series({'values': f'{test_name} Rsdl!$A$1:$A$5'})
        # chart.add_series({'values': f'{test_name} Rsdl!$B$1:$B$5'})
        for i in range(1, forecast_df_cols + 1):
            chart.add_series(
                {
                    "name": [f"{workbook_test_name} Predicted", 0, i, 0, i],
                    "categories": [
                        f"{workbook_test_name} Predicted",
                        1,
                        0,
                        len(forecast_df),
                        0,
                    ],
                    "values": [
                        f"{workbook_test_name} Predicted",
                        1,
                        i,
                        len(forecast_df),
                        i,
                    ],
                }
            )

        # Set chart title and labels
        chart.set_title({"name": "Predicted traffic plot"})
        chart.set_x_axis({"name": "Time"})
        chart.set_y_axis({"name": "Predicted AADT"})

        # Insert the chart into the worksheet
        worksheet.insert_chart("I2", chart)

        g_df.to_excel(
            writer, sheet_name=f"{workbook_test_name} regr inputs", index=True
        )
        # Move the pointer of the buffer to the beginning
        buffer.seek(0)

    return buffer


def reformat_excel(buffer, test_name):
    workbook_test_name = shorten_test_name(test_name)
    buffer.seek(0)  # Reset the buffer to the start before reading it

    # Load the workbook and select the desired worksheet
    wb = load_workbook(buffer)
    ws = wb["All regressions"]

    # Define styles for bold text and grey background
    bold_font = Font(bold=True)
    grey_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    # Find the row with "test_name" in column A
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value == test_name:
                row_number = cell.row

                # Apply formatting to the entire row
                for cell_in_row in ws[row_number]:
                    cell_in_row.font = bold_font
                    cell_in_row.fill = grey_fill

                # Save the modified workbook
                # wb.save(buffer)
                break
    # Format numbers to 3 decimal places in Column B (or any other numerical columns)
    for row in ws.iter_rows(
        min_row=2, min_col=2, max_col=22
    ):  # Adjust column range for other numerical columns
        for cell in row:
            cell.number_format = "0.000"  # 3 decimal places


    ws_coeff = wb[f"{workbook_test_name} Coeffs"]
    adjust_column_width(ws_coeff)
    # Format numbers to 3 decimal places in Column B (or any other numerical columns)
    for row in ws_coeff.iter_rows(
            min_row=2, min_col=2, max_col=2
    ):  # Adjust column range for other numerical columns
        for cell in row:
            cell.number_format = "0.000"  # 3 decimal places


    # Save the modified workbook to a new buffer
    new_buffer = io.BytesIO()
    wb.save(new_buffer)
    new_buffer.seek(0)  # Reset buffer pointer to the beginning after saving

    # Close the workbook to release resources
    wb.close()

    return new_buffer

def shorten_test_name(test_name):
    # avoid having too long worksheet names, which causes errors when saving workbooks (max chars 31)
    if len(test_name) > 20:
        shortened_test_name_elements = [x[:3] for x in test_name.split("-")]
        workbook_test_name = "".join(shortened_test_name_elements)
        if len(workbook_test_name) > 20:
            workbook_test_name = workbook_test_name[:15] + "+"
    else:
        workbook_test_name = test_name
    return workbook_test_name


def adjust_column_width(worksheet):
    # Loop over all columns in the worksheet
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', etc.)

        # Iterate over all cells in the column, including the header
        for cell in col:
            try:
                # Calculate the length of the cell's string representation
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Set the column width (adding a little extra for padding)
        worksheet.column_dimensions[column].width = max_length + 2  # Adding some padding
