import openpyxl


def create_input_template(client_name, project_name, y_variables, x_variables):
    # Define the path to the template file
    template_path = "data/utils/excel_template_v0.01.xlsx"

    # Load the template workbook
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Define the values to be updated
    file_name = f"{project_name} Regression Inputs"
    heading = "Regression Inputs"

    # Update the specified cells
    ws["B2"] = client_name
    ws["B3"] = project_name
    ws["B4"] = file_name
    ws["B6"] = heading

    # Define y variable names
    y_row = 10

    ws[f"D{y_row}"] = "Dependent Variables"
    ws[f"E{y_row}"] = "abs/pct"

    for key, item in y_variables.items():
        y_row += 1

        ws[f"D{y_row}"] = key
        ws[f"E{y_row}"] = item

    # Define x variable names
    x_row = y_row + 2

    ws[f"D{x_row}"] = "Independent Variables"
    ws[f"E{x_row}"] = "abs/pct"

    for key, item in x_variables.items():
        x_row += 1

        ws[f"D{x_row}"] = key
        ws[f"E{x_row}"] = item

    # Define the output path
    output_path = f"data/reg_input/{file_name}.xlsx"

    # Save the modified workbook
    wb.save(output_path)


# Example usage for development
if __name__ == "__main__":
    y_variables = {"Traffic (AADT)": "abs"}

    x_variables = {
        "GDP": "abs",
        "Unemployment": "pct",
        "Ramp Up": "abs",
        "Signage": "abs",
        "Cong_IC2": "abs",
        "Cong_A20": "abs",
    }

    create_input_template("AEDL", "A32 LV Traffic Forecast", y_variables, x_variables)
