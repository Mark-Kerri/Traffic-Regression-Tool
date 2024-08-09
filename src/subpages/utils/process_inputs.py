import os
import openpyxl
import pandas as pd
import streamlit as st
import plotly.express as px
import numpy as np
# Assumptions for the following script to work:
# Column D only has values for dependent and independent variables
# and no other content (otherwise any column D text will be assumed a new variable)
# Data for each variable found in Column D is being read from column G onwards
# until an empty cell is found (None)
# Output dataframe will name each variable before "Independent Variables" cell
# as X:{variable name} and every variable after "Independent Variables" cell as Y:{variable name}
# Building the index of the df assuming column G is empty (on all rows above data)
# and column G is the first time series column

def stringify(i:int = 0) -> str:
    # print(f'stringify prints {i}')
    """This function allows the slider to have string values rather than just numbers"""
    return st.session_state.df_index[i]

def stringify_g_df(i:int = 0) -> str:
    # print(f'stringify prints {i}')
    """This function allows the slider to have string values rather than just numbers"""
    # print(st.session_state.g_df_idx)
    return st.session_state.g_df_idx[i]
def spreadsheet_to_df(file_name):

    # delete below line to test pathing
    # file_path = os.path.join("data", "reg_input")
    # file_name = "Development Test Data Regression Inputs.xlsx"
    # full_path = os.path.join(file_path, file_name)
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    sheet = workbook.active

    # Read column names and row names so that the dataframe is filled
    df_cols = []
    dependent_flag = 1
    col = 4  # start from Column D for variable names
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is not None:
            # if sheet.cell(row=row, column=col).value != 'Dependent Variables'
            # and  sheet.cell(row=row, column=col).value != 'Independent Variables' :
            if sheet.cell(row=row, column=col).value == "Dependent Variables":
                dependent_flag = 1
                continue
            if sheet.cell(row=row, column=col).value == "Independent Variables":
                dependent_flag = 0
                continue
            if dependent_flag == 1:
                df_cols.append("y:" + sheet.cell(row=row, column=col).value)
            elif dependent_flag == 0:
                df_cols.append("x:" + sheet.cell(row=row, column=col).value)

    # Build dataframe by iterating over the columns for each variable
    df = pd.DataFrame(columns=df_cols)
    for c in df_cols:
        temp_list = []
        # print(c)
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=col).value == c[2:]:
                start_row = row
                break
        cell_names = sheet[start_row]
        for cell_obj in cell_names[6:]:
            try:
                val = float(str(cell_obj.value))
                temp_list.append(val)
            except ValueError:
                break
        df[c] = temp_list

    # Build the index of the df assuming column G is empty
    # and column G is the first time series column
    col = 7  # Start at Col G
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is not None:
            start_row = row
            # print(start_row)
            break
    idx_row = start_row + 2  # 2 rows below first non-empty cell

    temp_list = []
    cell_names = sheet[idx_row]
    # print(cell_names)
    for cell_obj in cell_names[(col - 1) :]:
        try:
            val = str(cell_obj.value)
            if val != "None":
                temp_list.append(val)
        except ValueError:
            break
    df_idx = temp_list
    df.index = df_idx

    # Read variable type/unit and build a dictionary
    var_dict = {}
    col = 5  # Start at Column E
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value == 'abs/pct' or (sheet.cell(row=row, column=col).value is None):
            continue
        var_dict[sheet.cell(row=row, column=col-1).value] = sheet.cell(row=row, column=col).value
    df_index = df.index

    return df,df_index,var_dict


def visualise_data(df, slider_value_start,slider_value_end):
#= st.session_state.slider_value_start,
#                    = st.session_state.slider_value_end):
    # print(f'df length : {len(df)}')
    # print(f'slider_value : {slider_value}')
    print(f'start value: {slider_value_start}, {stringify(slider_value_start)}')
    print(f'end value: {slider_value_end}, {stringify(slider_value_end)}')

    df = df[slider_value_start:slider_value_end+1]
    # df = df[:]
    print(slider_value_start)
    print(slider_value_end)

# print(f'df length : {len(df)}')
    df_idx = df.index
    df_x = pd.DataFrame(index=df_idx)
    df_y = pd.DataFrame(index=df_idx)

    # split x and y columns into separate dfs
    for col in df.columns:
        if col[0] == "x":
            df_x[col] = df[col]
        elif col[0] == "y":
            df_y[col] = df[col]
    fig = px.line(df, x=df.index, y=df.columns,title="Interactive chart of each variable over time")
    fig.update_layout(
        xaxis_title="Year",
        yaxis_title="Variable"
    )
    st.plotly_chart(fig)

    df_indexed = 100 * (df / df.iloc[0, :])
    fig = px.line(df_indexed, x=df_indexed.index, y=df_indexed.columns,
                  title="Interactive chart of each variable indexed to base-100 over time")
    fig.update_layout(
        xaxis_title="Year",
        yaxis_title="Indexed variable"
    )
    st.plotly_chart(fig)

def growth_df(df):
    # 4 for quarterly data
    prd = 4
    print(df.columns)
    print(st.session_state.var_dict)
    #identify columns of each type
    for df_col in df.columns:
        if st.session_state.var_dict[df_col[2:]] == 'abs':
            df['g: ' + df_col] = df[df_col].pct_change(periods=prd)+1 # different forms of this produce
            # same coefficient but different constant (e.g. pct_change + 1, or pct_change * 100, or pct_change only
        if st.session_state.var_dict[df_col[2:]] == 'pct_val_or_dummy':
            df['g: ' + df_col] = np.exp(df[df_col] - df[df_col].shift(prd))
        # e.g. real toll change:
        if st.session_state.var_dict[df_col[2:]] == 'pct_change':
            df['g: ' + df_col] = df[df_col] + 1


        g_cols = [c for c in df.columns if c[0:2]=='g:']
        # print(g_cols)
        g_df = df[g_cols].dropna(how='all')
        g_df_idx = g_df.index
        print('g_df_idx:', g_df.index)

    return g_df,g_df_idx
def growth_list(elements):
    return [f"g: {element}" for element in elements]
def show_df(df,x_sel,y_sel):
    # def quarter_to_datetime(quarter_str):
    #     year, quarter = quarter_str.split(' Q')
    #     month = (int(quarter) - 1) * 3 + 1  # January for Q1, April for Q2, etc.
    #     return pd.Timestamp(f"{year}-{month:02d}-01")
    # print('x_sel: ',x_sel)
    # print('y_sel: ',y_sel,'/n')
    filt_cols = []
    filt_cols.append(y_sel)
    for x in x_sel:
        filt_cols.append(x)

    st.dataframe(data=st.session_state.df[st.session_state.slider_value_start:st.session_state.slider_value_end+1][filt_cols])
    # df['date'] = df.index.map(quarter_to_datetime)


    # df['annual_growth'] = df.groupby(df['date'].dt.quarter)['traffic'].pct_change(periods=4) * 100
    # print(g_df.head())
