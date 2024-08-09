"""
Excel Input Template Generator Module.

This module provides functionality for generating and managing Excel input
templates for project data analysis. It uses the `openpyxl` library to create,
manipulate, and style Excel spreadsheets based on various project inputs and
timeline details.

Constants:
- TEMPLATE_PATH (str): Path to the Excel template file used as the basis for
  the generated input files.
- OUTPUT_DIR (str): Directory path where the output files will be saved.
- HEADER_STYLE_CELL (str): Cell reference for the header style used in the
  generated template.
- VALUE_STYLE_CELL (str): Cell reference for the value style used in the
  generated template.

Functions:
- copy_cell_style(source_cell, target_cell):
    Copies the style attributes from a source cell to a target cell, including
    font, border, fill, number format, alignment, and protection.

- generate_timeline(timeline_inputs):
    Generates a timeline based on input parameters such as timestep, start year,
    and end year. Returns a dictionary with years, steps, and combined timeline
    strings.

- create_input_template(name_variables, y_variables, x_variables,
    timeline_inputs, file_name, output_folder_path):
    Creates an Excel input template based on provided project details, variables,
    and timeline inputs. Handles exceptions related to file operations and
    input validation, saving the final template to the specified directory.

- insert_columns_with_style(ws, start_column, num_columns):
    Inserts columns into the specified worksheet and applies the style from an
    existing column to the new columns.

- update_basic_info(ws, client_name, project_name, file_name):
    Updates the basic project information (client name, project name, and file
    name) in the worksheet.

- add_variables_with_timeline(ws, variables, header_text, start_row, timelines):
    Adds variable names, types, and corresponding timeline data to the worksheet,
    including the header and value styling.

Exceptions:
- FileNotFoundError: Raised if the template file is not found.
- KeyError: Raised if a required key is missing in input dictionaries.
- TypeError: Raised if there is a type mismatch in the inputs.
- openpyxl.utils.exceptions.InvalidFileException: Raised if the Excel template
  is invalid.
- OSError: Raised if there is a problem with file I/O operations.

This module is designed to facilitate the creation of structured and well-styled
Excel templates for data analysis and project management, ensuring consistency
and accuracy in the generated reports.
"""

import os
from copy import copy
from calendar import month_abbr
import openpyxl

# Constants
TEMPLATE_PATH = "data/utils/excel_template_v0.01.xlsx"
OUTPUT_DIR = "data/reg_input"
HEADER_STYLE_CELL = "D13"
VALUE_STYLE_CELL = "D14"


def copy_cell_style(source_cell, target_cell):
    """
    Copy the style of a source cell to a target cell.

    Parameters:
        source_cell (openpyxl.cell.Cell): The cell from which to copy the style.
        target_cell (openpyxl.cell.Cell): The cell to which the style is copied.
    """
    try:
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
    except AttributeError as e:
        print(f"Warning: Unable to copy some cell properties. Error: {e}")


def generate_timeline(timeline_inputs):
    """
    Generate a timeline based on the provided inputs.

    Parameters:
        timeline_inputs (dict): A dictionary containing timeline parameters
                                such as 'Timestep', 'Start Year', 'End Year', etc.

    Returns:
        dict: A dictionary with keys 'years', 'steps', and 'combined',
              representing different components of the timeline.
    """
    timestep = timeline_inputs["Timestep"]
    start_year = timeline_inputs["Start Year"]
    start_step = timeline_inputs["Start Timestep"]
    end_year = timeline_inputs["End Year"]
    end_step = timeline_inputs["End Timestep"]

    # Validation
    if start_year > end_year:
        raise ValueError("Start Year must be less than or equal to End Year")

    if timestep not in ["Monthly", "Quarterly", "Yearly"]:
        raise ValueError("Timestep must be Monthly, Quarterly, or Yearly")

    if timestep == "Monthly" and (
        start_step < 1 or start_step > 12 or end_step < 1 or end_step > 12
    ):
        raise ValueError(
            "For Monthly timestep, Start and End Timestep must be between 1 and 12"
        )
    elif timestep == "Quarterly" and (
        start_step < 1 or start_step > 4 or end_step < 1 or end_step > 4
    ):
        raise ValueError(
            "For Quarterly timestep, Start and End Timestep must be between 1 and 4"
        )
    elif timestep == "Yearly" and (start_step != 1 or end_step != 1):
        raise ValueError("For Yearly timestep, Start and End Timestep must be 1")

    list_1, list_2, list_3 = [], [], []
    steps_per_year = (
        12 if timestep == "Monthly" else 4 if timestep == "Quarterly" else 1
    )

    for year in range(start_year, end_year + 1):
        start = start_step if year == start_year else 1
        end = end_step if year == end_year else steps_per_year

        for step in range(start, end + 1):
            list_1.append(year)

            if timestep == "Monthly":
                step_str = month_abbr[step]
            elif timestep == "Quarterly":
                step_str = f"Q{step}"
            else:  # Yearly
                step_str = ""

            list_2.append(step_str)
            list_3.append(f"{year} {step_str}".strip())

    timelines = {"years": list_1, "steps": list_2, "combined": list_3}
    return timelines


def create_input_template(
    name_variables,
    y_variables,
    x_variables,
    timeline_inputs,
    file_name,
    output_folder_path,
):
    """
    Create an Excel input template based on project details, variables, and timeline inputs.

    Parameters:
        name_variables (dict): Project and client names.
        y_variables (dict): Dependent variables and their types.
        x_variables (dict): Independent variables and their types.
        timeline_inputs (dict): Timeline details.
        file_name (str): Desired name for the output file.
        output_folder_path (str): Directory path where the file will be saved.

    Raises:
        FileNotFoundError: If the template file is not found.
        KeyError: If a required key is missing in input dictionaries.
        TypeError: If there is a type mismatch in the inputs.
        openpyxl.utils.exceptions.InvalidFileException: If the Excel template is invalid.
        OSError: If there is a problem with file I/O operations.
    """
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active

        client_name = name_variables["Client"]
        project_name = name_variables["Project"]
        update_basic_info(ws, client_name, project_name, file_name)

        timelines = generate_timeline(timeline_inputs)
        num_columns = len(timelines["combined"])
        insert_columns_with_style(ws, 7, num_columns)

        last_row = add_variables_with_timeline(
            ws, y_variables, "Dependent Variables", 13, timelines
        )
        add_variables_with_timeline(
            ws, x_variables, "Independent Variables", last_row + 5, timelines
        )

        output_path = os.path.join(output_folder_path, f"{file_name}.xlsx")
        wb.save(output_path)
        wb.close()

        print(f"Template created successfully: {output_path}")

    except FileNotFoundError:
        print(f"Template file not found at path: {TEMPLATE_PATH}")
    except KeyError as ke:
        print(f"Missing key in input data: {ke}")
    except TypeError as te:
        print(f"Type error encountered: {te}")
    except openpyxl.utils.exceptions.InvalidFileException as ife:
        print(f"Invalid file format encountered: {ife}")
    except OSError as oe:
        print(f"File system error when saving the file: {oe}")


def insert_columns_with_style(ws, start_column, num_columns):
    """
    Insert columns into the worksheet and applies the style from an existing column.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to modify.
        start_column (int): The column index where insertion starts.
        num_columns (int): Number of columns to insert.
    """
    max_row = ws.max_row
    ws.insert_cols(start_column, num_columns)

    for row in range(1, max_row + 1):
        source_cell = ws.cell(row=row, column=start_column + num_columns)
        for col in range(start_column, start_column + num_columns):
            target_cell = ws.cell(row=row, column=col)
            copy_cell_style(source_cell, target_cell)


def update_basic_info(ws, client_name, project_name, file_name):
    """
    Update basic project information in the worksheet.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to update.
        client_name (str): Name of the client.
        project_name (str): Name of the project.
        file_name (str): The name of the output file.
    """
    ws["B2"] = client_name
    ws["B3"] = project_name
    ws["B4"] = file_name
    ws["B6"] = "Regression Inputs"


def add_variables_with_timeline(ws, variables, header_text, start_row, timelines):
    """
    Add variables and their corresponding timeline to the worksheet.

    Parameters:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to update.
        variables (dict): A dictionary of variables and their types.
        header_text (str): The header text to display.
        start_row (int): The starting row for adding the variables.
        timelines (dict): The generated timeline data.

    Returns:
        int: The last row index after adding the variables.
    """
    ws[f"D{start_row}"] = header_text
    ws[f"E{start_row}"] = "abs/pct"

    header_style = ws[HEADER_STYLE_CELL]
    value_style = ws[VALUE_STYLE_CELL]
    year_step_style = ws["D12"]

    copy_cell_style(header_style, ws[f"D{start_row}"])
    copy_cell_style(header_style, ws[f"E{start_row}"])

    for col, value in enumerate(timelines["years"], start=7):
        cell = ws.cell(row=start_row - 2, column=col)
        cell.value = value
        copy_cell_style(year_step_style, cell)

    for col, value in enumerate(timelines["steps"], start=7):
        cell = ws.cell(row=start_row - 1, column=col)
        cell.value = value
        copy_cell_style(year_step_style, cell)

    for col, value in enumerate(timelines["combined"], start=7):
        cell = ws.cell(row=start_row, column=col)
        cell.value = value
        copy_cell_style(header_style, cell)

    row = start_row
    for key, item in variables.items():
        row += 1
        ws[f"D{row}"] = key
        ws[f"E{row}"] = item

        copy_cell_style(value_style, ws[f"D{row}"])
        copy_cell_style(value_style, ws[f"E{row}"])

        for col in range(7, 7 + len(timelines["combined"])):
            cell = ws.cell(row=row, column=col)
            copy_cell_style(value_style, cell)

    return row
